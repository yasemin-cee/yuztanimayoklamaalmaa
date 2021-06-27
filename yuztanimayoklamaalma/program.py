import os
from datetime import date
from datetime import datetime
from tkinter import *
from tkinter import messagebox
import time

import cv2
import numpy as np
import pandas as pd
import xlwt
from PIL import Image
from openpyxl import *

# arayüz oluşturuluyor.
ana_pencere = Tk()
ana_pencere.title("Yoklama Sistemi")
yazi_tipi = "Helvetica 18 bold"
ana_pencere.geometry("400x250")

# buton resimleri ekleniyor.
yoklama = PhotoImage(file=r"kontrol.png")
yoklamaResim = yoklama.subsample(2, 2)

newUserPhoto = PhotoImage(file=r"new_user.png")
yeniOgrenci = newUserPhoto.subsample(2, 2)

baslik = Label(ana_pencere, text="Yoklama Sistemi", font=yazi_tipi)
baslik.grid(column=1, row=1)
def YoklamaAlma():
    style_string = "font: bold on; borders: bottom dashed"
    style = xlwt.easyxf(style_string)
    style1 = xlwt.easyxf(num_format_str='D-MMM-YY')

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet('Gelenler', cell_overwrite_ok=True)
    bir_col = ws.col(0)
    bir_col.width = 256 * 20
    iki_col = ws.col(1)
    iki_col.width = 256 * 20
    uc_col = ws.col(2)
    uc_col.width = 256 * 20
    tarih_col = ws.col(4)
    tarih_col.width = 256 * 13

    df = pd.read_excel('sinif_listesi.xlsx')
    ad = df['Ad']
    soyad = df['Soyad']
    no = df['Numara']
    uzunluk = len(df)

    ws.write(0, 3, datetime.now(), style1)
    ws.write(0, 0, 'Numara', style=style)
    ws.write(0, 1, 'Ad', style=style)
    ws.write(0, 2, 'Soyad', style=style)
    for z in range(0, uzunluk):
        ws.write(z + 1, 0, str(no[z]))
        ws.write(z + 1, 1, ad[z])
        ws.write(z + 1, 2, soyad[z])
        ws.write(z + 1, 3, "-")

    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read('ogrenme.yml')
    cascadePath = "haarcascade_frontalface_default.xml"
    faceCascade = cv2.CascadeClassifier(cascadePath);

    cam = cv2.VideoCapture(0)
    font = cv2.FONT_HERSHEY_SIMPLEX
    while True:
        ret, im = cam.read()
        gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
        faces = faceCascade.detectMultiScale(gray, 1.2, 5)
        for (x, y, w, h) in faces:
            cv2.rectangle(im, (x, y), (x + w, y + h), (150, 0, 0), 2)
            Id, conf = recognizer.predict(gray[y:y + h, x:x + w])

            for i in range(0, uzunluk):
                if Id == no[i]:
                    ws.write(i + 1, 3, "+", style=style)

            cv2.putText(im, str(Id), (x + 5, y + h - 5), font, 0.5, (0, 0, 255), 1, cv2.LINE_AA)
        cv2.imshow('YUZ TANIMA UYGULAMASI', im);

        if (cv2.waitKey(1) == ord('q')):
            wb.save('Yoklama/gelenler ' + str(date.today()) + '.xls')
            break;
    cam.release()
    cv2.destroyAllWindows()


def YeniOgrenciKaydet():
    kayit_penceresi = Tk()
    # arkaplan rengi
    kayit_penceresi.configure(background='light green')
    # pencere başlığı
    kayit_penceresi.title("Kayıt Formu")
    # pencere boyutu
    kayit_penceresi.geometry("500x300")
    kayit_baslik = Label(kayit_penceresi, text="Yeni Ogrenci Kayıt Et", bg="light green", font=yazi_tipi)
    # etiketler olusturuluyor.
    ogrenci_no = Label(kayit_penceresi, text="Ogrenci No", bg="light green")
    ad = Label(kayit_penceresi, text="Ad", bg="light green")
    soyad = Label(kayit_penceresi, text="Soyad", bg="light green")
    # yerlesşimler yapılıyor. 
    kayit_baslik.grid(row=0, column=1)
    ogrenci_no.grid(row=1, column=0)
    ad.grid(row=2, column=0)
    soyad.grid(row=3, column=0)
    # giriş yapılacak alanlar ekleniyor.
    ogrenci_no_field = Entry(kayit_penceresi)
    ad_field = Entry(kayit_penceresi)
    soyad_field = Entry(kayit_penceresi)
    ogrenci_no_field.grid(row=1, column=1, ipadx="100")
    ad_field.grid(row=2, column=1, ipadx="100")
    soyad_field.grid(row=3, column=1, ipadx="100")
    wb = load_workbook('sinif_listesi.xlsx')
    # excele kaydetmek için bir çalışma sayfası ayrılıyor.
    sheet = wb.active

    def YuzAlma():
        try:
            cam = cv2.VideoCapture(0)
            detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
            Id = ogrenci_no_field.get()
            sampleNum = 0
            while (True):
                ret, img = cam.read()
                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                faces = detector.detectMultiScale(gray, 1.3, 5)
                for (x, y, w, h) in faces:
                    cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    sampleNum = sampleNum + 1
                    cv2.imwrite("yuzler/User." + Id + '.' + str(sampleNum) + ".jpg", gray[y:y + h, x:x + w])
                    cv2.imshow('YUZ TARAMA', img)

                if cv2.waitKey(100) & 0xFF == ord('q'):
                    break

                elif sampleNum > 80:
                    break

            cam.release()
            cv2.destroyAllWindows()
        except:
            print("Hata")
        else:
            recognizer = cv2.face.LBPHFaceRecognizer_create()
            detector = cv2.CascadeClassifier("haarcascade_frontalface_default.xml");

            def getImagesAndLabels(path):

                imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
                faceSamples = []
                Ids = []

                for imagePath in imagePaths:

                    if (os.path.split(imagePath)[-1].split(".")[-1] != 'jpg'):
                        continue

                    pilImage = Image.open(imagePath).convert('L')
                    imageNp = np.array(pilImage, 'uint8')
                    Id = int(os.path.split(imagePath)[-1].split(".")[1])
                    faces = detector.detectMultiScale(imageNp)

                    for (x, y, w, h) in faces:
                        faceSamples.append(imageNp[y:y + h, x:x + w])
                        Ids.append(Id)
                return faceSamples, Ids

            faces, Ids = getImagesAndLabels('yuzler')
            recognizer.train(faces, np.array(Ids))
            recognizer.save('ogrenme.yml')

    def clear():
        # metin girişleri yapılacak alanları temizleyen metot.
        ad_field.delete(0, END)
        soyad_field.delete(0, END)
        ogrenci_no_field.delete(0, END)

    def OgrenciyiKaydet():
        if ad_field.get() == "" or soyad_field.get() == "" or ogrenci_no_field.get() == "":
            print("Tüm alanlar doldurulmadı!")
            messagebox.showerror("Hata!", "Lütfen tüm alanları doldurun!")
        else:
            try:
                current_row = sheet.max_row
                current_column = sheet.max_column
                sheet.cell(row=current_row + 1, column=1).value = ogrenci_no_field.get()
                sheet.cell(row=current_row + 1, column=2).value = ad_field.get()
                sheet.cell(row=current_row + 1, column=3).value = soyad_field.get()
                wb.save('sinif_listesi.xlsx')
                ad_field.focus_set()

            except:
                messagebox.showerror("Hata", "Kayıt Başarısız!")
            else:
                # kayıt gerçekleşirse bir bilgi mesajı çıkıyor ve clear metodu çağırılarak alanlar temizleniyor.
                messagebox.showinfo("Kayıt Tamamlandı!", "Öğrenci Kayıt Edildi!")
                clear()


    kaydet_btn = Button(kayit_penceresi, text="Kaydet", fg="White", bg="Red", command=OgrenciyiKaydet)
    kaydet_btn.grid(row=8, column=0, ipadx="50", ipady="20")
    yuz_tara_btn = Button(kayit_penceresi, text="Yüz Tara", command=YuzAlma)
    yuz_tara_btn.grid(row=8, column=1, ipadx="50", ipady="20")


yeni_ogrenci_btn = Button(ana_pencere, text="Yeni Kayıt",font=yazi_tipi, image=yeniOgrenci, compound=TOP, bg="light green",
                          fg="blue", command=YeniOgrenciKaydet)
yeni_ogrenci_btn.grid(column=0, row=5)

yoklamaAl_btn = Button(ana_pencere, text="Yoklama Al", font=yazi_tipi, image=yoklamaResim, compound=TOP, bg="light green", fg="blue", command=YoklamaAlma)
yoklamaAl_btn.grid(column=1, row=5)

mainloop()
